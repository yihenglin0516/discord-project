import re
from numpy.lib.shape_base import tile
import openpyxl
import pandas as pd
import discord
import matplotlib.pyplot as plt


class ClassSchedule:
    def __init__(self):
        self.dic = {'星期一':'B','星期二':'C','星期三':'D','星期四':'E','星期五':'F','星期六':'G'}

    def create(self,username):
        self.schedule = openpyxl.Workbook()
        sheet = self.schedule.worksheets[0]
        sheet.title = '課表'

        row_range = sheet['B1':'G1'][0]
        for i in range(6):
            row_range[i].value = list(self.dic.keys())[i]

        time_list = ['時間','0 7:10~8:20','1 8:10~9:00','2 9:10~10:00','3 10:20~11:10','4 11:20~12:10','5 12:20~13:10','6 13:20~14:10',
        '7 14:20~15:10','8 15:30~16:20','9 16:30~17:20','10 17:30~18:20','A 18:25~19:15','B 19:20~20:10','C 20:15~21:05','D 21:10~22:00']
        column_range = sheet['A1':'A16']
        for i in range(16):
            column_range[i][0].value = time_list[i]

        sheet['A17'].value = '沒有課程時間的課'
        sheet['H1'].value = '目前學分'
        sheet['I1'].value = '0'
        self.schedule.save(username + 'class_schedule.xlsx')


    def add(self,CoursenameTeacher,hash,username):
        try:
            wb = openpyxl.load_workbook(username + 'class_schedule.xlsx')
            sheet = wb.worksheets[0]
        except FileNotFoundError:
            self.create(username)  #如果沒有課表就新增一個
            wb = openpyxl.load_workbook(username + 'class_schedule.xlsx')
            sheet = wb.worksheets[0]
        course_info = hash.search(CoursenameTeacher)  # attribute : Name , Teacher , Time , link
        name , teacher , when ,credit = course_info.Name , course_info.Teacher ,course_info.Time ,course_info.credit
        class_name = name + '('  + teacher + ')'
        sheet['I1'].value = str(int(sheet['I1'].value) + int(float(credit)))
        wb.save(username + 'class_schedule.xlsx')
        # print(sheet['I1'].value)
        if when == '' :  ##跑不進來這個loop
            sheet['B17'].value = sheet['B17'].value + name + '(' + teacher + ')' + '\n'  # 把沒有課程時間的都放在B17
            wb.save(username + 'class_schedule.xlsx')

        else:
            when  = re.sub(r"\(.*?\)|\{.*?\}|\[.*?\]", " ", when)   # 星期二3,4 星期三6,7
            class_time = when.split(' ')       # Ex : class_time = ['星期一3,4','星期二6,7']
            count = 0
            for i in class_time :
                temp = re.split(r'(\d+)', i) # ['星期二', '3', ',', '4', '']
                temp = temp[:-1] # 去掉最後那個空白
                for k in temp :
                    if k == ',' :
                        pass
                    elif k in self.dic:
                        add_date = k
                    else :
                        # wb = openpyxl.load_workbook('class_schedule.xlsx')
                        # sheet = wb.worksheets[0]
                        add_row = int(k) + 2   # Ex : 第3節課要加在第5個row
                        add_column = self.dic[add_date] # Ex : 星期二要加在第C欄
                        # class_name = name + '('  + teacher + ')'
                        if sheet[str(add_column) + str(add_row)].value == None:
                            sheet[str(add_column) + str(add_row)].value = class_name + '\n'
                            wb.save(username + 'class_schedule.xlsx')
                        elif class_name in sheet[str(add_column) + str(add_row)].value.split('\n')[:-1] :
                            return False
                        else:
                            if count == 0 :
                                deleted_class = str(sheet[str(add_column) + str(add_row)].value.split('\n')[:-1][0])
                                deleted_class = deleted_class.replace('(',' ').replace(')','')
                                self.delete(deleted_class,hash,username)
                                count +=1
                            wb = openpyxl.load_workbook(username + 'class_schedule.xlsx')
                            sheet = wb.worksheets[0]
                            print(class_name)
                            sheet[str(add_column) + str(add_row)].value = class_name + '\n'
                            wb.save(username + 'class_schedule.xlsx')

    def delete(self,CoursenameTeacher,hash,username):
        course_info = hash.search(CoursenameTeacher)  # attribute : Name , Teacher , Time , link
        name , teacher , when , credit = course_info.Name , course_info.Teacher ,course_info.Time , course_info.credit
        wb = openpyxl.load_workbook(username + 'class_schedule.xlsx')
        sheet = wb.worksheets[0]
        sheet['I1'].value = str(int(sheet['I1'].value) - int(float(credit)))
        # print(sheet['I1'].value)
        if when == '' :     #沒有上課時間的課程
            added_course = sheet['B17'].value.split('\n')
            added_course = added_course[:-1]
            delete_course = name + '(' + teacher + ')'
            if delete_course in added_course :
                added_course.remove(delete_course)
                sheet['B17'].value = '\n'.join(added_course) + '\n'
                wb.save(username + 'class_schedule.xlsx')

        else :
            when  = re.sub(r"\(.*?\)|\{.*?\}|\[.*?\]", " ", when)   # 星期二3,4 星期三6,7
            class_time = when.split(' ')       # Ex : class_time = ['星期一3,4','星期二6,7']
            for i in class_time :
                temp = re.split(r'(\d+)', i) # ['星期二', '3', ',', '4', '']
                temp = temp[:-1] # 去掉最後那個空白
                for k in temp :
                    if k == ',':
                        pass
                    elif k in self.dic:
                        delete_date = k
                    else:
                        delete_row = int(k) + 2   # Ex : 第3節課在第5個row
                        delete_column = self.dic[delete_date] # Ex : 星期二在第C欄
                        added_course = sheet[str(delete_column) + str(delete_row)].value.split('\n')
                        added_course = added_course[:-1] # ['資料結構(吳沛遠)','電子學(鍾孝文)']
                        delete_course = name + '(' + teacher + ')'
                        if delete_course in added_course :
                            added_course.remove(delete_course)
                            sheet[str(delete_column) + str(delete_row)].value = '' + '\n'
                            wb.save(username + 'class_schedule.xlsx')
                        else :
                            return False

    def show(self,username):
        chart = pd.read_excel(username + 'class_schedule.xlsx')
        chart = chart.fillna('-')

        embed = discord.Embed(title = username + "的課表")
        embed.add_field(name = "時間",value='\n'.join(chart["時間"].tolist()),inline=True)
        count = 0
        for date in self.dic:
            if count == 2:
                embed.add_field(name = "時間",value='\n'.join(chart["時間"].tolist()),inline=True)
                count = 0

            date_list = chart[date].tolist()
            for i in range(len(date_list)-1):
                date_list[i] = date_list[i].strip()
            embed.add_field(name = date,value='\n'.join(date_list),inline=True)
            count += 1
        return embed

    def show_credit(self,username):
        wb = openpyxl.load_workbook(username + 'class_schedule.xlsx')
        sheet = wb.worksheets[0]
        embed = discord.Embed(title = username + "目前的學分")
        embed.add_field(name="學分",value=str(sheet['I1'].value),inline=True)
        return embed
